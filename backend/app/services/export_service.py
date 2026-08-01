from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

TRIP_TYPE_LABEL = {"directo": "Directo", "indirecto": "Indirecto"}

FE_HEADERS = [
    "Código de Camión",
    "Placa",
    "Destino Inicial",
    "Destino Final",
    "Tipo de Viaje",
    "Monto",
    "Marca",
    "Foto VIN",
]
FE_COLUMN_WIDTHS = [18, 12, 18, 18, 14, 12, 18, 14]


def build_fe_workbook(rows: list[dict]) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "FE"
    ws.append(FE_HEADERS)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for row in rows:
        ws.append(
            [
                row.get("truck_code") or "",
                row.get("truck_plate") or "",
                row.get("origin") or "",
                row.get("destination") or "",
                TRIP_TYPE_LABEL.get(row.get("trip_type"), ""),
                float(row["amount"]) if row.get("amount") is not None else "",
                row.get("brand") or "",
                "Ver foto" if row.get("vin_photo_url") else "",
            ]
        )
        vin_photo_url = row.get("vin_photo_url")
        if vin_photo_url:
            cell = ws.cell(row=ws.max_row, column=len(FE_HEADERS))
            cell.hyperlink = vin_photo_url
            cell.style = "Hyperlink"

    for i, width in enumerate(FE_COLUMN_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
